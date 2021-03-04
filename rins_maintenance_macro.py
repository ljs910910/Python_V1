import pywinauto as pwa
import pyautogui
import os
import time
import warnings
import errno
import datetime
from PIL import Image, ImageGrab
import pytesseract
import re
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
import openpyxl
import cv2

warnings.simplefilter('ignore', category=UserWarning)

def setFocus(title_reg):
    app = pwa.application.Application()
    t = title_reg
    print('find title : ' + str(title_reg))
    try:
        handle = pwa.findwindows.find_windows(title_re=t)[0]
        app.connect(handle=handle)
        print('title: ' + str(t) + 'handle: ' + str(handle) + 'Setted')
    except:
        print('No title exist on window')
    window = app.window(handle=handle)
    try:
        window.set_focus()
    except Exception as e:
        print('[error]setFocuse : ' + str(e))
    return window

def IOI():
    t = u'STGClient.*'
    return setFocus(t)

if __name__ == "__main__":
    IOI()
    rins_find_btn = pyautogui.locateOnScreen('rins_find_btn.png')
    pyautogui.moveTo(rins_find_btn)
    pyautogui.moveRel(90, 0)
    pyautogui.click()
    pyautogui.hotkey('enter')
    time.sleep(1)

def TACS_Control():
    rins_find_btn1 = pyautogui.locateOnScreen('rins_find_btn1.png')
    pyautogui.moveTo(rins_find_btn1)
    pyautogui.moveRel(122, 0)
    pyautogui.click()
    time.sleep(1)
    pyautogui.typewrite('rins-v-prox01')
    pyautogui.hotkey('enter')
    time.sleep(1)
    pyautogui.moveRel(0, 70)
    time.sleep(1)
    pyautogui.doubleClick()
    time.sleep(3)
    pyautogui.hotkey('win', 'up')
    time.sleep(1)
    pyautogui.click(100,100, button='right')
    time.sleep(1)
    rins_find_btn2 = pyautogui.locateOnScreen('change_settings.png')
    pyautogui.moveTo(rins_find_btn2)
    pyautogui.click()
    time.sleep(1)
    rins_find_btn3 = pyautogui.locateOnScreen('appearance.png')
    pyautogui.moveTo(rins_find_btn3)
    pyautogui.click()
    time.sleep(1)
    rins_find_btn4 = pyautogui.locateOnScreen('change.png')
    pyautogui.moveTo(rins_find_btn4)
    pyautogui.click()
    time.sleep(2)
    rins_find_btn5 = pyautogui.locateOnScreen('size.png')
    pyautogui.moveTo(rins_find_btn5)
    pyautogui.moveRel(0, 60)
    time.sleep(1)
    pyautogui.click()
    pyautogui.hotkey('enter')
    time.sleep(1)
    rins_find_btn6 = pyautogui.locateOnScreen('apply.png')
    pyautogui.moveTo(rins_find_btn6)
    pyautogui.click()
    time.sleep(1)

def Capture():
    img = ImageGrab.grab()
    img.save('./rins_image/' + str(rins_server_list1) + '.png')
    time.sleep(2)

def DB_Capture():
    img = ImageGrab.grab()
    img.save('./rins_image/' + str(rins_server_list2) + '.png')
    time.sleep(2)

def now_date():
    now = datetime.datetime.now()
    now_dt = now.strftime('%Y-%m-%d')
    return now_dt

def img_modify():
    # hostname_img = cv2.imread('./rins_image/' + str(rins_server_list1) + '.png', cv2.IMREAD_COLOR)
    hostname_img = cv2.imread('./rins_image/' + str(rins_server_list1) + '.png')
    img_cut = hostname_img[23:100, 0:300].copy()
    # img_cut1 = cv2.cvtColor(img_cut, cv2.COLOR_BGR2HLS_FULL)
    height, width = img_cut.shape[:2]
    resize_img = cv2.resize(img_cut, (2 * width, 2 * height), interpolation=cv2.INTER_LANCZOS4)
    img_result = cv2.imwrite('./rins_image/' + str(rins_server_list3) + '.png', 255 - resize_img)
    time.sleep(1)
    return img_result

def OCR():
    OCR = pytesseract.image_to_string(Image.open('./rins_image/' + str(rins_server_list3) + '.png'))
    print(OCR)
    Catch = re.search(str(rins_server_list3), OCR)
    print(Catch)
    return Catch

def rins_pw():
    pyautogui.typewrite('ssh rinsop@' + rins_server_list)
    pyautogui.hotkey('enter')
    time.sleep(1)
    pyautogui.typewrite('passwd')
    pyautogui.hotkey('enter')
    time.sleep(1)
    pyautogui.typewrite('clear')
    pyautogui.hotkey('enter')

try:
    if not (os.path.isdir('rins_image')):
        os.makedirs(os.path.join('rins_image'))
except OSError as e:
    if e.errno != errno.EEXIST:
        print('fail')
        exit()

f1 = open('rins_server_list.txt', 'r')
f2 = open('rins_server_list1.txt', 'r')
f3 = open('rins_server_list2.txt', 'r')
f4 = open('rins_server_list3.txt', 'r')

cnt = 0
while cnt < 1:
    cnt += 1
    rins_server_list1 = f2.readline().rstrip()
    rins_server_list3 = f4.readline().rstrip()
    print(rins_server_list1)
    TACS_Control()
    pyautogui.typewrite('clear')
    pyautogui.hotkey('enter')
    time.sleep(1)
    Capture()
    img_modify()
    if OCR() is not None:
        print('login success')
        time.sleep(1)
        pyautogui.typewrite('tail -f /data/logs/WS_SPPS_J_0.log.' + now_date())
        pyautogui.hotkey('enter')
        time.sleep(10)
        pyautogui.hotkey('ctrl', 'c')
        time.sleep(2)
        pyautogui.typewrite('tail -f /data/logs/WS_SPPS_J_1.log.' + now_date())
        pyautogui.hotkey('enter')
        time.sleep(10)
        pyautogui.hotkey('ctrl', 'c')
        time.sleep(1)
        pyautogui.typewrite('date; mpstat | tail -1 | awk \'{print 100-$NF}\'; free -m; df -h')
        pyautogui.hotkey('enter')
        time.sleep(1)
        Capture()
    else:
        print('not match, error')
        break
    if cnt == 1:
        while cnt < 2:
            cnt += 1
            rins_server_list = f1.readline().rstrip()
            rins_server_list1 = f2.readline().rstrip()
            rins_server_list3 = f4.readline().rstrip()
            print(rins_server_list1)
            rins_pw()
            time.sleep(1)
            Capture()
            img_modify()
            if OCR() is not None:
                print('login success')
                time.sleep(1)
                pyautogui.typewrite('tail -f /data/logs/SPPS_J_0.log.' + now_date())
                pyautogui.hotkey('enter')
                time.sleep(10)
                pyautogui.hotkey('ctrl', 'c')
                time.sleep(2)
                pyautogui.typewrite('tail -f /data/logs/SPPS_J_1.log.' + now_date())
                pyautogui.hotkey('enter')
                time.sleep(10)
                pyautogui.hotkey('ctrl', 'c')
                time.sleep(1)
                pyautogui.typewrite('date; mpstat | tail -1 | awk \'{print 100-$NF}\'; free -m; df -h')
                pyautogui.hotkey('enter')
                time.sleep(1)
                Capture()
            else:
                print('not match, error')
                break
            if cnt == 2:
                while cnt < 7:
                    cnt += 1
                    rins_server_list = f1.readline().rstrip()
                    rins_server_list1 = f2.readline().rstrip()
                    rins_server_list3 = f4.readline().rstrip()
                    print(rins_server_list1)
                    rins_pw()
                    time.sleep(1)
                    Capture()
                    img_modify()
                    if OCR() is not None:
                        print('login success')
                        time.sleep(1)
                        pyautogui.typewrite('tail -f /data/logs/kafka/kafka.log')
                        pyautogui.hotkey('enter')
                        time.sleep(10)
                        pyautogui.hotkey('ctrl', 'c')
                        time.sleep(1)
                        pyautogui.typewrite('ls -alh /data/logs/kafka/kafka-error.log')
                        pyautogui.hotkey('enter')
                        time.sleep(1)
                        pyautogui.typewrite('date; mpstat | tail -1 | awk \'{print 100-$NF}\'; free -m; df -h')
                        pyautogui.hotkey('enter')
                        time.sleep(1)
                        Capture()
                    else:
                        print('not match, error')
                        break
                    if cnt == 7:
                        while cnt < 9:
                            cnt += 1
                            rins_server_list = f1.readline().rstrip()
                            rins_server_list1 = f2.readline().rstrip()
                            rins_server_list3 = f4.readline().rstrip()
                            print(rins_server_list1)
                            rins_pw()
                            time.sleep(1)
                            Capture()
                            img_modify()
                            if OCR() is not None:
                                print('login success')
                                time.sleep(1)
                                pyautogui.typewrite('vi /data/logs/batch/batch.log')
                                pyautogui.hotkey('enter')
                                time.sleep(1)
                                pyautogui.typewrite('gg')
                                time.sleep(4)
                                for batch_cnt in range(0,9):
                                    batch_cnt = pyautogui.hotkey('ctrl', 'f')
                                    time.sleep(4)
                                pyautogui.hotkey('shift', ':')
                                time.sleep(1)
                                pyautogui.typewrite('q!')
                                pyautogui.hotkey('enter')
                                time.sleep(1)
                                pyautogui.typewrite('cat /data/logs/batch/batch-error.log')
                                pyautogui.hotkey('enter')
                                time.sleep(2)
                                pyautogui.typewrite('date; mpstat | tail -1 | awk \'{print 100-$NF}\'; free -m; df -h')
                                pyautogui.hotkey('enter')
                                time.sleep(1)
                                Capture()
                            else:
                                print('not match, error')
                                break
                            if cnt == 9:
                                while cnt < 10:
                                    cnt += 1
                                    rins_server_list = f1.readline().rstrip()
                                    rins_server_list1 = f2.readline().rstrip()
                                    rins_server_list3 = f4.readline().rstrip()
                                    print(rins_server_list1)
                                    rins_pw()
                                    time.sleep(1)
                                    Capture()
                                    img_modify()
                                    if OCR() is not None:
                                        print('login success')
                                        time.sleep(1)
                                        pyautogui.typewrite('tail -f /data/logs/spcs/spcs.' + now_date() + '.log')
                                        pyautogui.hotkey('enter')
                                        time.sleep(10)
                                        pyautogui.hotkey('ctrl', 'c')
                                        time.sleep(1)
                                        pyautogui.typewrite('date; mpstat | tail -1 | awk \'{print 100-$NF}\'; free -m; df -h')
                                        pyautogui.hotkey('enter')
                                        time.sleep(1)
                                        Capture()
                                    else:
                                        print('not match, error')
                                        break
                                    if cnt == 10:
                                        while cnt < 11:
                                            cnt += 1
                                            rins_server_list = f1.readline().rstrip()
                                            rins_server_list1 = f2.readline().rstrip()
                                            rins_server_list3 = f4.readline().rstrip()
                                            print(rins_server_list1)
                                            rins_pw()
                                            time.sleep(1)
                                            Capture()
                                            img_modify()
                                            if OCR() is not None:
                                                print('login success')
                                                time.sleep(1)
                                                pyautogui.typewrite('tail -f /data/logs/push/push.' + now_date() + '.log')
                                                pyautogui.hotkey('enter')
                                                time.sleep(10)
                                                pyautogui.hotkey('ctrl', 'c')
                                                time.sleep(1)
                                                pyautogui.typewrite('date; mpstat | tail -1 | awk \'{print 100-$NF}\'; free -m; df -h')
                                                pyautogui.hotkey('enter')
                                                time.sleep(1)
                                                Capture()
                                            else:
                                                print('not match, error')
                                                break
                                            if cnt == 11:
                                                while cnt < 12:
                                                    cnt += 1
                                                    rins_server_list = f1.readline().rstrip()
                                                    rins_server_list1 = f2.readline().rstrip()
                                                    rins_server_list3 = f4.readline().rstrip()
                                                    print(rins_server_list1)
                                                    rins_pw()
                                                    time.sleep(1)
                                                    Capture()
                                                    img_modify()
                                                    if OCR() is not None:
                                                        print('login success')
                                                        time.sleep(1)
                                                        pyautogui.typewrite('tail -f /data/logs/spls/spls.log')
                                                        pyautogui.hotkey('enter')
                                                        time.sleep(10)
                                                        pyautogui.hotkey('ctrl', 'c')
                                                        time.sleep(1)
                                                        pyautogui.typewrite('date; mpstat | tail -1 | awk \'{print 100-$NF}\'; free -m; df -h')
                                                        pyautogui.hotkey('enter')
                                                        time.sleep(1)
                                                        Capture()
                                                    else:
                                                        print('not match, error')
                                                        break
                                                    if cnt == 12:
                                                        while cnt < 21:
                                                            cnt += 1
                                                            rins_server_list = f1.readline().rstrip()
                                                            rins_server_list1 = f2.readline().rstrip()
                                                            rins_server_list2 = f3.readline().rstrip()
                                                            rins_server_list3 = f4.readline().rstrip()
                                                            print(rins_server_list1)
                                                            rins_pw()
                                                            time.sleep(1)
                                                            Capture()
                                                            img_modify()
                                                            if OCR() is not None:
                                                                print('login success')
                                                                time.sleep(1)
                                                                pyautogui.typewrite('su -')
                                                                pyautogui.hotkey('enter')
                                                                time.sleep(1)
                                                                pyautogui.typewrite('passwd')
                                                                pyautogui.hotkey('enter')
                                                                time.sleep(1)
                                                                pyautogui.typewrite('su - mysql')
                                                                pyautogui.hotkey('enter')
                                                                time.sleep(1)
                                                                pyautogui.typewrite('sql')
                                                                pyautogui.hotkey('enter')
                                                                time.sleep(1)
                                                                pyautogui.typewrite('show slave status\G;')
                                                                time.sleep(1)
                                                                pyautogui.hotkey('enter')
                                                                time.sleep(1)
                                                                DB_Capture()
                                                                pyautogui.typewrite('quit')
                                                                pyautogui.hotkey('enter')
                                                                time.sleep(1)
                                                                pyautogui.typewrite('exit')
                                                                pyautogui.hotkey('enter')
                                                                time.sleep(1)
                                                                pyautogui.typewrite('exit')
                                                                pyautogui.hotkey('enter')
                                                                time.sleep(1)
                                                                pyautogui.typewrite('date; mpstat | tail -1 | awk \'{print 100-$NF}\'; free -m; df -h')
                                                                pyautogui.hotkey('enter')
                                                                time.sleep(1)
                                                                Capture()
                                                            else:
                                                                print('not match, error')
                                                                break
                                                            if cnt == 21:
                                                                break
f1.close()
f2.close()
f3.close()
f4.close()

load_wb = openpyxl.Workbook()
def create_excel():
    load_ws = load_wb.active
    load_wb.create_sheet(index=1, title='total_sheet')

if __name__ == "__main__":
    create_excel()

def create_date():
    cell = ['G1', 'F1', 'E1', 'D1', 'C1', 'B1', 'A1']
    for i in range(-7, 0):
        day_delta = datetime.timedelta(days=1)
        start_date = datetime.date.today()
        a_week_ago = start_date + i * datetime.timedelta(1)
        result = a_week_ago.strftime('%Y-%m-%d')
        load_ws2 = load_wb['total_sheet']
        load_ws2[cell.pop()] = result

        # 금 ~ 목 API 이력관리 엑셀 파일 다운
        driver.get('http://172.16.1.1/spa/admin/services/apiLog/excelDownload.do?searchResult=&apiGroupCd=&searchStDt=' + result + '&searchStHH=00&searchStMM=00&searchEdDt=' + result + '&searchEdHH=23&searchEdMM=59&logContent=')

# 상용 cms 로그인
while True:
    try:
        driver = webdriver.Chrome(r'C:\Users\webiznet\Desktop\기타\chromedriver.exe')
        driver.get('http://172.16.1.1/spa/account/login.do')
        driver.find_element_by_name('j_username').send_keys('admin')
        driver.find_element_by_name('j_password').send_keys('passwd')
        driver.find_element_by_xpath("""//*[@id="loginForm"]/fieldset/p/input""").click()

        for k in range(-7, 0):  # 금 ~ 목 시트 생성 후, 일별 데이터 저장
            day_delta = datetime.timedelta(days=1)
            start_date = datetime.date.today()
            a_week_ago = start_date + k * datetime.timedelta(1)
            result = a_week_ago.strftime('%Y%m%d')
            load_wb.create_sheet(index=1, title=result)
            load_ws = load_wb[result]
            driver.get('http://172.16.1.1/spa/admin/mng/pushmng/timetraffic.ajax?searchDt=' + str(result))
            html = driver.page_source
            soup = BeautifulSoup(html, 'html.parser')
            time.sleep(1)
            for a_tag in soup.findAll('td', class_='ac'):
                result1 = (a_tag.get_text().strip().replace('\n', '').replace('\t', ''))
                load_ws = load_wb[result]
                load_ws.append([result1])

        cell1 = ['G2', 'F2', 'E2', 'D2', 'C2', 'B2', 'A2']
        for j in range(-7, 0):  # 최댓값 비교
            day_delta = datetime.timedelta(days=1)
            start_date = datetime.date.today()
            a_week_ago = start_date + j * datetime.timedelta(1)
            result = a_week_ago.strftime('%Y%m%d')
            load_ws = load_wb[result]
            if load_ws['A6'].value < load_ws['A9'].value:
                temp3 = load_ws['A9'].value
            else:
                temp3 = load_ws['A6'].value
            for q in range(9, 76, 3):
                n3 = load_ws['A' + str(q)].value
                if temp3 < n3:
                    temp3 = n3
                else:
                    pass
            load_ws2 = load_wb['total_sheet']
            load_ws2[cell1.pop()] = temp3
            print(result, temp3)

    except NoSuchElementException as e:
        print('retry', e)
        continue
        
    break

create_date()
load_wb.remove(load_wb['Sheet'])
load_wb.save('crawlling.xlsx')
