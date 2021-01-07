from bs4 import BeautifulSoup
from selenium import webdriver
import datetime
import time
import openpyxl

load_wb = openpyxl.Workbook()

def create_excel():
    load_ws = load_wb.active
    load_wb.create_sheet(index=1, title='total_sheet')

if __name__ == "__main__":
    create_excel()

def create_date():
    cell = ['G1', 'F1', 'E1', 'D1', 'C1', 'B1', 'A1']
    for i in range(-7, 0):
        day_delta = datetime.timedelta(days=1)
        start_date = datetime.date.today()
        a_week_ago = start_date + i * datetime.timedelta(1)
        result = a_week_ago.strftime('%Y-%m-%d')
        load_ws2 = load_wb['total_sheet']
        load_ws2[cell.pop()] = result

        # 금 ~ 목 API 이력관리 엑셀 파일 다운
        driver.get('http://172.16.42.50/spa/admin/services/apiLog/excelDownload.do?searchResult=&apiGroupCd=&searchStDt=' + result + '&searchStHH=00&searchStMM=00&searchEdDt=' + result + '&searchEdHH=23&searchEdMM=59&logContent=')
        
# 상용 cms 로그인
driver = webdriver.Chrome(r'C:\Users\webiznet\Desktop\기타\chromedriver.exe')
driver.get('http://172.16.42.50/spa/account/login.do')
driver.find_element_by_name('j_username').send_keys('rins_admin')
driver.find_element_by_name('j_password').send_keys('ipaleldj1!')
driver.find_element_by_xpath("""//*[@id="loginForm"]/fieldset/p/input""").click()

for k in range(-7, 0):  # 금 ~ 목 시트 생성 후, 일별 데이터 저장
    day_delta = datetime.timedelta(days=1)
    start_date = datetime.date.today()
    a_week_ago = start_date + k * datetime.timedelta(1)
    result = a_week_ago.strftime('%Y%m%d')
    load_wb.create_sheet(index=1, title=result)
    load_ws = load_wb[result]
    driver.get('http://172.16.42.50/spa/admin/mng/pushmng/timetraffic.ajax?searchDt=' + str(result))
    html = driver.page_source
    soup = BeautifulSoup(html, 'html.parser')
    time.sleep(1)
    for a_tag in soup.findAll('td', class_='ac'):
        result1 = (a_tag.get_text().strip().replace('\n', '').replace('\t', ''))
        load_ws = load_wb[result]
        load_ws.append([result1])

cell1 = ['G2', 'F2', 'E2', 'D2', 'C2', 'B2', 'A2']
for j in range(-7, 0):  # 최댓값 비교
    day_delta = datetime.timedelta(days=1)
    start_date = datetime.date.today()
    a_week_ago = start_date + j * datetime.timedelta(1)
    result = a_week_ago.strftime('%Y%m%d')
    load_ws = load_wb[result]
    if load_ws['A6'].value < load_ws['A9'].value:
        temp3 = load_ws['A9'].value
    else:
        temp3 = load_ws['A6'].value
    for q in range(9, 76, 3):
        n3 = load_ws['A' + str(q)].value
        if temp3 < n3:
            temp3 = n3
        else: pass
    load_ws2 = load_wb['total_sheet']
    load_ws2[cell1.pop()] = temp3
    print(result, temp3)

create_date()
load_wb.remove(load_wb['Sheet'])
load_wb.save('crawlling.xlsx')
